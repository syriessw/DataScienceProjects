'''
Basic code to use Selenium to download an Excel from a Holiday management site.
The Excel will be saved to a designated location.
Thereafter, ExcelWriter will be used to make some changes to remove private information.
This is the first part in a series of automation to send a holiday management file
to assigned person for resource allocation on a weekly basis.
'''

import glob
import time
import os
import sys
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from datetime import datetime
from datetime import timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils.cell import get_column_letter

import requests
import json
import io

import os.path

#for SSL cert
import ssl
import certifi
from urllib.request import urlopen

#set a watcher for when file is downloading before proceeding to close
def download_wait(directory, timeout, nfiles=None):
    '''
    Wait for downloads to finish with a specified timeout

    Args
    ----
    directory : str
        The path to the folder where the files will be downloaded.
    timeout : int
        How many seconds to wait until timing out.
    nfiles : int, defaults to None
        If provided, also wait for the expected number of files
    '''
    seconds = 0
    dl_wait = True
    while dl_wait and seconds < timeout:
        time.sleep(1)
        dl_wait = False
        files = os.listdir(directory)
        if nfiles and len(files) != nfiles:
            dl_wait = True
        
        for fname in files:
            if fname.endswith('.crdownload'):
                dl_wait = True
        
        seconds += 1
    return seconds

'''
Section for downloading from TMF
'''

### This is for personal information
### Can be place into a yaml file instead
myCompanyCode = ''
myEeId = ''
myPassword = ''

#Set the download directory
temp_directory = ''
# chrome_options = Options()

chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": os.getcwd() + "\Files\\",
    "download.prompt_for_download": False,
    "profile.default_content_settings.popups": 0,
    "download.directory_upgrade": True,
    "safebrowsing_for_trusted_sources_enabled": False,
    "safebrowsing.enabled": False
})
chromeLocalStatePrefs = {'browser.enabled_labs_experiments': ['download-bubble@2', 'download-bubble-v2@2']}
chrome_options.add_experimental_option('localState', chromeLocalStatePrefs)

#Set SSL verify as false
os.environ['WDM_SSL_VERIFY'] = '0'

sys.path.insert(0, '/usr/lib/chromium-browser/chromedriver')

#Had error where chromedriver doesn't match, so added new one that covers that
#driver = webdriver.Chrome(options= chrome_options)
chrome_install = ChromeDriverManager().install()
folder = os.path.dirname(chrome_install)
chromedriver_path = os.path.join(folder, "chromedriver.exe")
chrome_service = webdriver.ChromeService(chromedriver_path)
driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
wait = WebDriverWait(driver,20)

#Save this current window handle
parent_handle = driver.current_window_handle

#Open the page
driver.get('https://app2.unit4hrms.com/prosoft/web')
#implicit wait
driver.implicitly_wait(10) #seconds

LoggedIn = len(driver.find_elements(By.ID, 'ctl00_cph1_lblName_eLeave'))
tries = 0

while LoggedIn < 1:

    #Identify the company box
    co_box = driver.find_element(By.ID, 'txtClientCode')
    #send company information
    co_box.send_keys(Keys.CONTROL + "a")
    co_box.send_keys(Keys.DELETE)
    co_box.send_keys(myCompanyCode)
    driver.implicitly_wait(3) #seconds

    #Identify the user box
    id_box = driver.find_element(By.ID, 'txtUserID')
    #send user information
    id_box.send_keys(Keys.CONTROL + "a")
    id_box.send_keys(Keys.DELETE)
    id_box.send_keys(myEeId)
    driver.implicitly_wait(3) #seconds

    #identiy the password box
    password_box = driver.find_element(By.ID, 'txtPassword')
    #send password information
    password_box.send_keys(Keys.CONTROL + "a")
    password_box.send_keys(Keys.DELETE)
    password_box.send_keys(myPassword)
    driver.implicitly_wait(3) #seconds

    #Click the button to login
    driver.find_element(By.NAME, 'btnSignIn').click()

    LoggedIn = len(driver.find_elements(By.ID, 'ctl00_cph1_lblName_eLeave'))
    print('Still trying to log in...')

    if(tries > 5):
        break

    tries = tries+1

print("Logged in successfully!")

#Enter the Leave section
driver.find_element(By.ID, 'ctl00_cph1_lblName_eLeave').click()

driver.implicitly_wait(5) #seconds

action = ActionChains(driver)

#Find the bar menu to hover over
leave_MenuButton = driver.find_element(By.XPATH, '//*[@id="ctl00_ucHeader1_radm1"]/ul/li[3]/a/span')
action.move_to_element(leave_MenuButton).perform()

driver.implicitly_wait(5) #seconds

#find the Leave Reports links
leave_MenuSecondary = driver.find_element(By.XPATH, '//*[@id="ctl00_ucHeader1_radm1"]/ul/li[3]/div/ul/li/div/div/table/tbody/tr/td[3]/table/tbody/tr[2]/td/a')

#click to enter the Leave Section
#leave_MenuSecondary.click()
#action.move_to_element(leave_MenuSecondary).perform()
driver.get("HTML OF BUTTON TO CLICK")

driver.implicitly_wait(5) #seconds

# A new window will open after clicking the button
# It will take a while to load, hence the implict wait and checking for action taken as the file downloads
#Because there is a new window, we will close this new window
#As there is no registered action in selenium, we will wait for 2nd window to load before closing it
def windowCheckClose():
    wait.until(EC.number_of_windows_to_be(2))
    all_han = driver.window_handles

    for window_handle in all_han:
        download_wait(os.getcwd(), 8, 1)

        if(window_handle != parent_handle):
            driver.switch_to.window(window_handle)
            driver.close()
            break
    
    #switch back to original window
    driver.switch_to.window(parent_handle)

'''
Download the monthly file and change the name
'''

# Check if current date is Monday Day 0
# If it is, then take days as of last Friday instead
now = datetime.now()

if now.weekday() == 0:
    start_date = now - timedelta(days=3)
else:
    start_date = now + timedelta(days = 3) #Get Current sending date + 3 days (for Monday)
start_date = datetime.now() + timedelta(days = 3)
start_date_Format = start_date.strftime("%d-%m-%Y")
print(start_date)
#Get end date + 1 month
one_mon_rel = relativedelta(months=1)
end_date = start_date + one_mon_rel
end_date_Format = end_date.strftime('%d-%m-%Y')
print(start_date_Format + " to " + end_date_Format)

#Select start date, clear, and enter variable
startDate_box = driver.find_element(By.ID, 'ctl00_cph1_ucDateStartEnd_dpDateStart_input')
#Clear the box of existing date
startDate_box.clear()
#Enter in the start Date
startDate_box.send_keys(start_date_Format)

#Select end date, clear and enter variable
endDate_box = driver.find_element(By.ID, 'ctl00_cph1_ucDateStartEnd_dpDateEnd_input')
#Clear the box of existing date
endDate_box.clear()
#Enter in the End Date
endDate_box.send_keys(end_date_Format)

#Download the Reports file
driver.find_element(By.NAME, 'ctl00$cph1$btnExportToExcelWorkbook').click()

windowCheckClose()
print("One Month File downloaded")

#Rename the monthly file to make changes for it

#get the list of files
list_of_files = glob.glob(os.path.join(os.getcwd(), "Files", "*.xlsx"))
#Get latest file after downloading
latest_file = max(list_of_files, key=os.path.getctime)
#currentTimeStamp = datetime.now().strftime('%Y%m%d')
startDateTimeStamp = start_date.strftime("%Y%m%d")
print("Now renaming " + latest_file)

outdir = r'.\\Files'
filename = 'SG_LeaveRecordsReportSingle_' + startDateTimeStamp + '.xlsx'
dir_filename = os.path.join(outdir, filename)
if(os.path.exists(dir_filename)):
    os.remove(dir_filename)
os.rename(latest_file, dir_filename)
print("One Month File renamed")

#Close the Chrome browser
driver.quit()

'''
Files has been downloaded
This section is for processing the downloaded files
'''

#get the list of files
list_of_files = glob.glob(os.path.join(os.getcwd(), "Files", "*.xlsx"))
#Get latest file after downloading
latest_file = max(list_of_files, key=os.path.getctime)
currentTimeStamp = datetime.now().strftime('%Y%m%d')
print("Now processing " + latest_file)

finaldir = r'.\\Files\Processed'

#check if Files directory exists
if not os.path.exists(outdir):
    os.mkdir(outdir)

fulldirectory = os.path.join(finaldir,filename)


#Check if file has been created before. If not, we create
if not os.path.exists(fulldirectory):
    #Using Openpyxl to drop a couple of columns for privacy reasons

    wb = load_workbook(filename=dir_filename)
    sheet = wb.active
    #Delete EE number col
    sheet.delete_cols(idx=1, amount = 1)
    #Delete some cols
    sheet.delete_cols(idx=4, amount = 3)
    sheet.delete_cols(idx=11, amount = 25)

    #make a table
    table = Table(displayName="HolidayList", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    sheet.add_table(table)

    wb.save(filename=fulldirectory)
    wb.close()
    print('New copy of One Month saved. Existing Workbook Closed')
else:
    print("One Month File" + filename + " has been processed before")
